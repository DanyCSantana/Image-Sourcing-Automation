#!/usr/bin/env python
# coding: utf-8

# # Source Movies Images (Posters and Stills)

# In[ ]:


import re
import shutil
import logging
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from typing import Optional, Dict, List
import textwrap

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

# ---------------------------------
# Requirements: pandas, openpyxl, tqdm
# ---------------------------------

# ---------------------------------
# Configuration - Adjustable as needed
# ---------------------------------
EXCEL_FILE = Path("Images tracker.xlsx")  # Default Excel file path
SOURCE_DIR = Path("Movies")                # Base directory for source images
DOWNLOAD_BASE_DIR = Path.home() / "Downloads"  # Default downloads folder

# ---------------------------------
# Logger Setup
# ---------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)


# ---------------------------------
# Utility Functions
# ---------------------------------
def clean_string(s: str) -> str:
    """Remove non-alphanumeric characters and lowercase the string."""
    return re.sub(r'\W+', '', s).lower() if s else ''


# ---------------------------------
# ImageManager Class
# ---------------------------------
class ImageManager:
    """Manage image search and copy operations."""

    def __init__(self, source_dir: Path, download_dir: Path):
        """Initialize with source and download directories."""
        self.source_dir = source_dir
        self.download_dir = download_dir
        self.copy_errors: List[str] = []

    def find_image(self, movie_title: str, distributor: str, folder_type: str) -> Optional[Path]:
        """Find an image matching movie title in distributor folder."""
        distributor_folder = self.source_dir / distributor
        if not distributor_folder.exists():
            logger.warning(f"Distributor folder not found: {distributor_folder}")
            return None

        poster_folders = [f for f in distributor_folder.iterdir() if 'Poster' in f.name and f.is_dir()]
        if len(poster_folders) == 1 and 'Horizontal Posters' in poster_folders[0].name:
            target_folder = poster_folders[0]
        else:
            target_folder = distributor_folder / folder_type

        if not target_folder.exists():
            logger.warning(f"Target folder does not exist: {target_folder}")
            return None

        matching_files = []
        for file in target_folder.rglob("*"):
            if file.is_file() and clean_string(movie_title) in clean_string(file.name):
                matching_files.append(file)

        for priority_tag in ["(1)", "(2)"]:
            for f in matching_files:
                if priority_tag in f.name:
                    return f

        if matching_files:
            return matching_files[0]

        return None

    def find_image_by_title_only(self, title: str, distributor: str, folder_type: str) -> Optional[Path]:
        """Find image by title only without season/episode."""
        dist_folder = self.source_dir / distributor
        if not dist_folder.exists():
            return None

        target_folder = dist_folder / folder_type
        if not target_folder.exists():
            return None

        for file in target_folder.rglob("*"):
            if file.is_file() and clean_string(title) in clean_string(file.name):
                return file

        return None

    def copy_image(self, src_path: Path, distributor: str, folder_type: str) -> Optional[str]:
        """Copy image to download directory, create folders if needed."""
        try:
            dest_folder = self.download_dir / distributor / folder_type
            dest_folder.mkdir(parents=True, exist_ok=True)
            dest_path = dest_folder / src_path.name
            shutil.copy(src_path, dest_path)
            logger.info(f"Copied image: {src_path} -> {dest_path}")
            return src_path.name
        except PermissionError as e:
            logger.error(f"Permission denied copying {src_path}: {e}")
            self.copy_errors.append(str(src_path))
            return None
        except Exception as e:
            logger.error(f"Error copying {src_path}: {e}")
            self.copy_errors.append(str(src_path))
            return None


# ---------------------------------
# ExcelHandler Class
# ---------------------------------
class ExcelHandler:
    """
    Handle Excel workbook operations for updating image data.
    """

    def __init__(self, excel_path: Path):
        """Load workbook and select the 'Movies' sheet."""
        self.excel_path = excel_path
        self.workbook = load_workbook(excel_path)
        self.sheet = self.workbook['Movies']

    def update_images(self, movie_data: Dict[str, Dict[str, Optional[str]]]) -> None:
        """
        Update the Excel sheet with poster and still image filenames.
        """
        sheet_title_map = {}
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row):
            cell = row[1]  # Column B (index 1)
            title = cell.value
            if title:
                cleaned = clean_string(str(title))
                sheet_title_map[cleaned] = cell

        for movie_title, images in movie_data.items():
            cleaned_title = clean_string(movie_title)
            if cleaned_title in sheet_title_map:
                cell = sheet_title_map[cleaned_title]
                if images.get("Poster"):
                    cell.offset(0, 2).value = images["Poster"]  # Column D
                if images.get("Still"):
                    cell.offset(0, 3).value = images["Still"]  # Column E

        self.workbook.save(self.excel_path)
        logger.info("Excel file updated with image filenames.")


# ---------------------------------
# Main Process Function
# ---------------------------------
def process_images(
    excel_file: Path,
    source_dir: Path,
    download_base_dir: Path
) -> None:
    """
    Process images: search, copy, update Excel and generate report.
    """
    today_str = datetime.now().strftime("%d%b")
    download_dir = download_base_dir / f"Images sourced {today_str}"
    download_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"Download directory set to: {download_dir}")

    distributor_contact_df = pd.read_excel(excel_file, sheet_name='Distributor_contact')
    df_movies = pd.read_excel(excel_file, sheet_name='Movies')

    img_manager = ImageManager(source_dir, download_dir)
    excel_handler = ExcelHandler(excel_file)

    movie_data: Dict[str, Dict[str, Optional[str]]] = {}
    not_found: Dict[str, List[str]] = defaultdict(list)

    logger.info("Starting image search...")

    for _, row in tqdm(df_movies.iterrows(), total=len(df_movies), desc="Processing movies"):
        title = row['Title']
        distributor = row['Distributor']

        title_str = str(title).strip() if pd.notna(title) else None
        distributor_str = str(distributor).strip() if pd.notna(distributor) else None

        if not title_str or not distributor_str:
            not_found[distributor_str or "UNDEFINED DISTRIBUTOR"].append(title_str or "Undefined title")
            continue

        poster_path = img_manager.find_image(title_str, distributor_str, "Poster")
        if not poster_path:
            poster_path = img_manager.find_image_by_title_only(title_str, distributor_str, "Poster")

        still_path = img_manager.find_image(title_str, distributor_str, "Still")
        if not still_path:
            still_path = img_manager.find_image_by_title_only(title_str, distributor_str, "Still")

        poster_name = img_manager.copy_image(poster_path, distributor_str, "Poster") if poster_path else None
        still_name = img_manager.copy_image(still_path, distributor_str, "Still") if still_path else None

        movie_data[title_str] = {"Poster": poster_name, "Still": still_name}

        if not poster_name and not still_name:
            not_found[distributor_str].append(title_str)

    excel_handler.update_images(movie_data)
    generate_report(not_found, download_dir, distributor_contact_df, img_manager.copy_errors)


# ---------------------------------
# Report Generator
# ---------------------------------
def generate_report(
    not_found: Dict[str, List[str]],
    download_dir: Path,
    distributor_contact_df: pd.DataFrame,
    copy_errors: List[str]
) -> None:
    """
    Generate a report txt listing missing images and copy errors.
    """
    total_missing = sum(len(titles) for titles in not_found.values())

    lines: List[str] = []
    if total_missing > 0:
        logger.warning(f"Images not found for {total_missing} titles.")
        lines.append(f"Images were not found for {total_missing} titles:\n")

        for distributor, titles in not_found.items():
            lines.append(f"DISTRIBUTOR {distributor.upper()}:")
            for title in titles:
                lines.append(f" - {title}")

            contact_info = distributor_contact_df[distributor_contact_df['Distributor'] == distributor]
            if not contact_info.empty:
                contact_name = contact_info.iloc[0]['Contact Name(s)']
                email = contact_info.iloc[0]['Email(s)']

                message = textwrap.dedent(f"""                    Hi {contact_name},

                    I hope this message finds you well.
                    Could you please assist with the poster and still images for the titles listed below?

                    {"".join(f"- {t}\n" for t in titles)}

                    These assets will be featured onboard the Condor 0225 update.

                    Thank you in advance.

                    Best regards,
                """)

                lines.append(f"\nEmail draft for {distributor}:")
                lines.append(f"To: {email}")
                lines.append("Message:")
                lines.append(message)
            else:
                lines.append(f"No contact information found for {distributor}")
    else:
        logger.info("All images were successfully found.")
        lines.append("All images were successfully found.")

    if copy_errors:
        logger.error(f"Errors copying {len(copy_errors)} files:")
        lines.append("\nErrors copying these files:")
        for err_file in copy_errors:
            lines.append(f"- {err_file}")

    report_path = download_dir / "image_search_report.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    logger.info(f"Process completed. Report saved to {report_path}")


# -------

